#!/usr/bin/env python3
import json, sys, copy
from pathlib import Path
from datetime import datetime, date, timedelta
try:
    from openpyxl import load_workbook
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--quiet"])
    from openpyxl import load_workbook

XLSX_PATH = Path("Resultado Gerencial 2026.xlsx")
OUT_PATH  = Path("docs/data.json")
MESES_NOMES = {1:"Jan",2:"Fev",3:"Mar",4:"Abr",5:"Mai",6:"Jun",7:"Jul",8:"Ago",9:"Set",10:"Out",11:"Nov",12:"Dez"}

def safe(v):
    if v is None: return 0.0
    try: return float(v)
    except: return 0.0

def parse_dre(wb, aba):
    if aba not in wb.sheetnames: return []
    ws = wb[aba]
    rows = list(ws.iter_rows(values_only=True))
    date_row = rows[3]
    meses = []
    for ci in range(1, len(date_row)):
        v = date_row[ci]
        if isinstance(v, (datetime, date)):
            meses.append({"ci":ci,"mes":MESES_NOMES.get(v.month,"?"),"mes_num":v.month,"ano":v.year,"order":v.year*100+v.month})
    LABEL_MAP = {
        "receita bruta de vendas":"rBruta","(-) devoluções e cancelamentos":"devolucoes",
        "(-) impostos sobre vendas":"impostos","receita líquida":"rLiquida",
        "(-) custo de mercadoria vendida (cmv)":"cmv","= lucro bruto":"lucroBruto",
        "(-) despesas de pessoal":"dPessoal","(-) despesas de marketing":"dMarketing",
        "(-) taxas de plataforma":"dTaxas","(-) despesas financeiras / financiamentos":"dFinanc",
        "(-) outras despesas operacionais":"dOutras","ebitda":"ebitda",
        "(-) depreciação / amortização":"depreciacao","ebit":"ebit","lucro líquido":"lucroLiquido",
    }
    row_data = {}
    i = 0
    while i < len(rows):
        row = rows[i]
        lbl = str(row[1]).strip().lower() if row[1] else ""
        if lbl in LABEL_MAP and i+1 < len(rows):
            row_data[LABEL_MAP[lbl]] = rows[i+1]; i += 2
        else: i += 1
    result = []
    for m in meses:
        ci = m["ci"]
        rec = {"mes":m["mes"],"mes_num":m["mes_num"],"ano":m["ano"],"order":m["order"]}
        for fname in ["rBruta","devolucoes","impostos","rLiquida","cmv","lucroBruto",
                      "dPessoal","dMarketing","dTaxas","dFinanc","dOutras","ebitda","depreciacao","ebit","lucroLiquido"]:
            rec[fname] = safe(row_data[fname][ci]) if fname in row_data else 0.0
        result.append(rec)
    return result

def parse_divida_fornecedores(wb):
    aba = "Divida Fornecedores"
    if aba not in wb.sheetnames: return {}
    ws = wb[aba]
    rows = list(ws.iter_rows(values_only=True))
    result = {"pagas":{"VIP":{},"VIDAL":{},"GRUPO":{}},"a_pagar":{"VIP":{},"GRUPO":{}}}
    for row in rows:
        for emp,ci_mes,ci_val in [("VIP",1,2),("VIDAL",5,6)]:
            v_mes = row[ci_mes] if len(row)>ci_mes else None
            v_val = row[ci_val] if len(row)>ci_val else None
            if isinstance(v_mes,(datetime,date)):
                mn = MESES_NOMES.get(v_mes.month,"?")
                val = safe(v_val)
                if val > 0: result["pagas"][emp][mn] = val
        v_mes = row[10] if len(row)>10 else None
        v_val = row[11] if len(row)>11 else None
        if isinstance(v_mes,(datetime,date)):
            mn = MESES_NOMES.get(v_mes.month,"?")
            val = safe(v_val)
            if val > 0: result["a_pagar"]["VIP"][mn] = val
    todos = set(list(result["pagas"]["VIP"].keys())+list(result["pagas"]["VIDAL"].keys()))
    for mn in todos:
        result["pagas"]["GRUPO"][mn] = result["pagas"]["VIP"].get(mn,0)+result["pagas"]["VIDAL"].get(mn,0)
    result["a_pagar"]["GRUPO"] = dict(result["a_pagar"]["VIP"])
    return result

def parse_custos(wb):
    if "Custos" not in wb.sheetnames: return {}
    ws = wb["Custos"]
    rows = list(ws.iter_rows(values_only=True))
    fixo=[]; variavel=[]; tf=0; tv=0
    for row in rows[3:]:
        lf = str(row[1]).strip() if row[1] else ""
        vf = safe(row[2]) if len(row)>2 else 0
        lv = str(row[5]).strip() if len(row)>5 and row[5] else ""
        vv = safe(row[6]) if len(row)>6 else 0
        if lf and lf!="ORIGEM" and vf>0: fixo.append({"label":lf,"valor":vf}); tf+=vf
        if lv and lv!="ORIGEM" and vv>0: variavel.append({"label":lv,"valor":vv}); tv+=vv
    return {"fixo":fixo,"variavel":variavel,"total_fixo":tf,"total_variavel":tv,"total":tf+tv}

def parse_plano(wb, aba_key):
    aba_map = {"VIP":"PLANO ORÇAMENTARIO - VIP","VIDAL":"PLANO ORÇAMENTARIO - VIDAL","V3":"PLANO ORÇAMENTARIO - V3"}
    aba = aba_map.get(aba_key,"")
    real_aba = next((s for s in wb.sheetnames if s.upper().replace("Ç","C").replace("Ã","A")==aba.upper().replace("Ç","C").replace("Ã","A")),None)
    if not real_aba: return []
    ws = wb[real_aba]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows)<3: return []
    pr_row,date_row = rows[0],rows[1]
    col_map = {}
    for ci in range(2,len(date_row)):
        v = date_row[ci]
        if isinstance(v,(datetime,date)):
            nome=MESES_NOMES.get(v.month,"?"); tipo_raw=str(pr_row[ci]).strip().lower() if pr_row[ci] else ""
            tipo="previsto" if "prev" in tipo_raw else "realizado"
            col_map[ci]={"mes":nome,"mes_num":v.month,"ano":v.year,"tipo":tipo,"order":v.year*100+v.month}
    LABEL_MAP={"receita bruta":"rBruta","(-) despesas do negocio":"despTotal","(-) despesas do negócio":"despTotal",
        "resultado":"resultado","resultado ":"resultado",
        "vendas mercado livre  - vip mx":"mlVip","vendas amazon - vip mx":"amazonVip",
        "vendas shopee  - vip mx":"shopeeVip","vendas loja fisica":"lojaFisica",
        "vendas mercado livre - vidal":"mlVidal","vendas shopee - vidal":"shopeeVidal",
        "vendas mercado livre":"ml","vendas shopee":"shopee","vendas amazon":"amazon"}
    row_data={}
    for row in rows[2:]:
        lbl=str(row[0]).strip().lower() if row[0] else ""
        if lbl in LABEL_MAP: row_data[LABEL_MAP[lbl]]=row
    meses_data={}
    for ci,info in col_map.items():
        key=(info["ano"],info["mes_num"])
        if key not in meses_data:
            meses_data[key]={"mes":info["mes"],"ano":info["ano"],"mes_num":info["mes_num"],"order":info["order"],
                "rPrev":0,"rReal":0,"dPrev":0,"dReal":0,"resPrev":0,"resReal":0,"canais":{}}
        suf="Prev" if info["tipo"]=="previsto" else "Real"
        for fname,rkey in [("r","rBruta"),("d","despTotal"),("res","resultado")]:
            if rkey in row_data: meses_data[key][fname+suf]+=safe(row_data[rkey][ci])
        for ck,cname in [("mlVip","Mercado Livre"),("ml","Mercado Livre"),("mlVidal","Mercado Livre"),
            ("shopeeVip","Shopee"),("shopee","Shopee"),("shopeeVidal","Shopee"),
            ("amazonVip","Amazon"),("amazon","Amazon"),("lojaFisica","Loja Fisica")]:
            if ck in row_data:
                v=safe(row_data[ck][ci])
                if v==0: continue
                if cname not in meses_data[key]["canais"]: meses_data[key]["canais"][cname]={"prev":0,"real":0}
                if info["tipo"]=="previsto": meses_data[key]["canais"][cname]["prev"]+=v
                else: meses_data[key]["canais"][cname]["real"]+=v
    result=sorted(meses_data.values(),key=lambda x:x["order"])
    com_real=[m for m in result if m["rReal"]>0]
    if com_real:
        ultimo=com_real[-1]["order"]
        return [m for m in result if m["order"]<=ultimo]
    return []

def parse_orcamento_2027(wb):
    aba = "Orçamento 2027"
    if aba not in wb.sheetnames: return []
    ws = wb[aba]
    rows = list(ws.iter_rows(values_only=True))
    date_row = rows[1]
    meses = []
    for ci in range(2, len(date_row)):
        v = date_row[ci]
        if isinstance(v,(datetime,date)):
            meses.append({"ci":ci,"mes":MESES_NOMES.get(v.month,"?"),"mes_num":v.month,"ano":v.year,"order":v.year*100+v.month})
    LABEL_MAP = {
        "receita bruta":"rBruta","vendas mercado livre":"mlVendas","vendas amazon":"amazonVendas",
        "vendas shopee":"shopeeVendas","vendas loja fisica":"lojaVendas","rendimentos":"rendimentos",
        "(-) despesas do negócio":"despTotal","despesas de pessoal":"dPessoal","prolabore":"dProlabore",
        "ceo":"dCeo","consultorias e acessorias":"dConsultoria","despesas gerais":"dGerais",
        "tecnologia":"dTecnologia","despesas de materia prima":"dMateria","despesas de vendas":"dVendas",
        "despesas de marketing":"dMarketing","impostos":"dImpostos","despesas financeiras":"dFinanceiras",
        "taxas plataforma":"dTaxas","cancelamentos e remmbolso":"dCancelamentos",
        "financiamentos":"dFinanciamentos","financiamentos ":"dFinanciamentos",
        "investimentos":"dInvestimentos","resultado":"resultado","resultado ":"resultado",
        "variavel ceo":"varCeo","resultado real":"resultadoReal",
    }
    row_data={}
    for row in rows:
        lbl=str(row[0]).strip().lower() if row[0] else ""
        if lbl in LABEL_MAP: row_data[LABEL_MAP[lbl]]=row
    result=[]
    for m in meses:
        ci=m["ci"]
        rec={"mes":m["mes"],"mes_num":m["mes_num"],"ano":m["ano"],"order":m["order"]}
        for fname in ["rBruta","mlVendas","amazonVendas","shopeeVendas","lojaVendas","rendimentos",
                      "despTotal","dPessoal","dProlabore","dCeo","dConsultoria","dGerais","dTecnologia",
                      "dMateria","dVendas","dMarketing","dImpostos","dFinanceiras","dTaxas",
                      "dCancelamentos","dFinanciamentos","dInvestimentos","resultado","varCeo","resultadoReal"]:
            rec[fname]=safe(row_data[fname][ci]) if fname in row_data else 0.0
        rec["canais"]={"Mercado Livre":rec["mlVendas"],"Amazon":rec["amazonVendas"],
                       "Shopee":rec["shopeeVendas"],"Loja Física":rec["lojaVendas"]}
        result.append(rec)
    return result

def build_json():
    if not XLSX_PATH.exists(): print(f"ERRO: {XLSX_PATH} nao encontrado."); sys.exit(1)
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    print(f"Abas: {len(wb.sheetnames)}")
    colors={"VIP":{"color":"#00C9A7","accent":"#00856E","label":"VIP MX"},
            "VIDAL":{"color":"#6C63FF","accent":"#4B44CC","label":"VIDAL"},
            "V3":{"color":"#FF6B6B","accent":"#CC4444","label":"V3"},
            "GRUPO":{"color":"#F5A623","accent":"#C47D0E","label":"Grupo VIP"}}
    dre={}
    for emp,aba in [("VIP","DRE - VIP"),("VIDAL","DRE - VIDAL"),("V3","DRE - V3")]:
        months=parse_dre(wb,aba); dre[emp]={**colors[emp],"months":months}
        print(f"  DRE {emp}: {len(months)} meses")
    grupo_dre={}
    for emp in ["VIP","VIDAL","V3"]:
        for m in dre[emp]["months"]:
            k=m["order"]
            if k not in grupo_dre: grupo_dre[k]=copy.deepcopy(m)
            else:
                for f in ["rBruta","devolucoes","impostos","rLiquida","cmv","lucroBruto",
                          "dPessoal","dMarketing","dTaxas","dFinanc","dOutras","ebitda","depreciacao","ebit","lucroLiquido"]:
                    grupo_dre[k][f]=grupo_dre[k].get(f,0)+m.get(f,0)
    dre["GRUPO"]={**colors["GRUPO"],"months":sorted(grupo_dre.values(),key=lambda x:x["order"])}
    print(f"  DRE GRUPO: {len(dre['GRUPO']['months'])} meses")
    plano={}
    for emp in ["VIP","VIDAL","V3"]:
        months=parse_plano(wb,emp); plano[emp]={**colors[emp],"months":months}
        print(f"  Plano {emp}: {len(months)} meses")
    all_keys={}
    for emp in ["VIP","VIDAL","V3"]:
        for m in plano[emp]["months"]:
            k=m["order"]
            if k not in all_keys: all_keys[k]=copy.deepcopy(m); all_keys[k]["canais"]={}
            else:
                for f in ["rPrev","rReal","dPrev","dReal","resPrev","resReal"]:
                    all_keys[k][f]=all_keys[k].get(f,0)+m.get(f,0)
            for cname,cv in m.get("canais",{}).items():
                if cname not in all_keys[k]["canais"]: all_keys[k]["canais"][cname]={"prev":0,"real":0}
                all_keys[k]["canais"][cname]["prev"]+=cv.get("prev",0)
                all_keys[k]["canais"][cname]["real"]+=cv.get("real",0)
    plano["GRUPO"]={**colors["GRUPO"],"months":sorted(all_keys.values(),key=lambda x:x["order"])}
    divida=parse_divida_fornecedores(wb)
    custos=parse_custos(wb)
    print(f"  Custos fixo: {len(custos.get('fixo',[]))} itens | variavel: {len(custos.get('variavel',[]))} itens")
    orcamento_2027=parse_orcamento_2027(wb)
    print(f"  Orçamento 2027: {len(orcamento_2027)} meses")
    OUT_PATH.parent.mkdir(parents=True,exist_ok=True)
    out={"gerado_em":datetime.now().strftime("%d/%m/%Y %H:%M"),"fonte":XLSX_PATH.name,
         "dre":dre,"plano":plano,"divida":divida,"custos":custos,"orcamento_2027":orcamento_2027}
    with open(OUT_PATH,"w",encoding="utf-8") as f: json.dump(out,f,ensure_ascii=False,indent=2)
    print(f"OK {OUT_PATH} gerado ({OUT_PATH.stat().st_size} bytes)")

if __name__ == "__main__": build_json()
